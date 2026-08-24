from openpyxl import load_workbook
from openpyxl.utils import range_boundaries
wb = load_workbook("MarkList.xlsx")
ws = wb.active
cell_range = input("Enter the cell range you want: ")
min_col, min_row, max_col, max_row = range_boundaries(cell_range)

#Finding all matching cells for a value we want:
Item = input("Enter the item you want to search: ")
found = False
count = 0
for row in range(min_row, max_row+1):
    for col in range(min_col, max_col+1):
        if str(ws.cell(row=row, column=col).value) == Item:
            print(f"{Item}","is found at",ws.cell(row,col).coordinate)
            found =True
            count = count+1
print("Total matches:",count)
if not found:
    print("Not found")

    #Output:
    # Enter the cell range you want: A1:F11
    # Enter the item you want to search: 90    
    # 90 is found at E2
    # 90 is found at F9
    # 90 is found at D11
    # Total matches: 3