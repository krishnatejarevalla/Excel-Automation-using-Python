from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
wb = load_workbook("MarkList.xlsx")
ws = wb.active

#Find the column number of each subject
for col in range(1, ws.max_column+1):
    column_letter = ws.cell(row= 1, column= col).column_letter
    Indexes = column_index_from_string(column_letter)
    print(ws.cell(row= 1, column = col).value+"=",Indexes)
    #Output:
    # Student_ID= 1
    # Name= 2
    # Python= 3
    # SQL= 4
    # Excel= 5
    # PowerBI= 6
