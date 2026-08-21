from openpyxl import load_workbook
from openpyxl.utils import get_column_interval
from openpyxl.utils import column_index_from_string, get_column_letter
wb = load_workbook("MarkList.xlsx")
ws = wb.active

#Write code to find the columns between B and F
print(get_column_interval("B","F"))
        #Output:
        # ['B', 'C', 'D', 'E', 'F']

#Printing column names btw C and F from column letters
columns = get_column_interval("C","F")
for column in columns:
    print(ws[column+"1"].value)
    #Output:
    # Python
    # SQL
    # Excel
    # PowerBI

#Printing column names dynamically
starting_column = input("Enter starting column letter: ")
ending_column = get_column_letter(ws.max_column)
column_range = get_column_interval(starting_column, ending_column)
for column in column_range:
    print(ws[column+"1"].value)
    #Output:
    # Enter starting column letter: B
    # Name
    # Python
    # SQL
    # Excel
    # PowerBI
