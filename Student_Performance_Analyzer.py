# ============================================================
#          Student Performance Analyzer
# ============================================================
# Description:
#This Project helps the teachers to get all the progress of a specific 
#student using their Name or ID
# ============================================================


from openpyxl import load_workbook
from openpyxl import Workbook

# STEP 1 - Load the Excel Workbook

wb = load_workbook("MarkList.xlsx")
ws = wb.active
print("=" * 60)
print("           STUDENT PERFORMANCE ANALYZER")
print("=" * 60)
print("Workbook loaded successfully!")
print("Worksheet:", ws.title)

#Finding total no of rows and columns that have data
print("Rows:", ws.max_row)
print("Columns:", ws.max_column)

        # OUTPUT:
        # Rows: 11
        # Columns: 6

#Printing Column Headers for Teacher's Reference
print("\nColumn Headers:")
for col in range(1, ws.max_column + 1):
    print(col, ws.cell(row=1, column=col).value)

        # OUTPUT:
        # Column Headers:
        # 1 Student_ID
        # 2 Name
        # 3 Python
        # 4 SQL
        # 5 Excel
        # 6 PowerBI

# STEP 2 - Identify Subjects Automatically

def get_subject_columns(ws):
    #Creating an empty dictionary
    subject_columns = {}
    for col in range(1, ws.max_column + 1):
        header = ws.cell(row=1, column=col).value
        if header not in ["Student_ID", "Name"]:
            #Storing Subject names and their column nums as key-value pairs in dict
            subject_columns[header] = col

    return subject_columns
subjects = get_subject_columns(ws)

#Printing the subjects-column numbers from dictionary
print("\nSubjects:")
for subject, column in subjects.items():
    print(subject, "-> Column", column)

        # OUTPUT:
        # Subjects:
        # Python -> Column 3
        # SQL -> Column 4
        # Excel -> Column 5
        # PowerBI -> Column 6

# STEP 3 - Find Student
def find_student(ws, search_value):
    for row in range(2, ws.max_row + 1):
        student_id = ws.cell(row=row, column=1).value
        student_name = ws.cell(row=row, column=2).value
        if str(student_id) == search_value or str(student_name).lower() == search_value.lower():
            return row

    return None


# STEP 4 - Get Student Subject Marks
def get_student_marks(ws, student_row, subjects):
    student_marks = {}
    for subject, column in subjects.items():
        marks = ws.cell(row=student_row, column=column).value
        student_marks[subject] = marks

    return student_marks   

# STEP 5 - Calculate Student Performance

def calculate_performance(student_marks):
    total = sum(student_marks.values())
    subject_count = len(student_marks)
    max_marks = subject_count * 100
    percentage = (total / max_marks) * 100
    average = total / subject_count
    return total, average, percentage

# STEP 6 - Analyze Student Performance

def analyze_student_performance(student_marks):

    highest_subject = max(student_marks, key=student_marks.get)
    lowest_subject = min(student_marks, key=student_marks.get)

    highest_marks = student_marks[highest_subject]
    lowest_marks = student_marks[lowest_subject]

    if highest_marks >= 90:
        performance_level = "Excellent"
    elif highest_marks >= 75:
        performance_level = "Good"
    elif highest_marks >= 50:
        performance_level = "Average"
    else:
        performance_level = "Needs Improvement"

    return (
        highest_subject,
        highest_marks,
        lowest_subject,
        lowest_marks,
        performance_level
    )

# STEP 7 - Calculate Student Grade

def calculate_grade(percentage):

    if percentage >= 90:
        return "A+"
    elif percentage >= 80:
        return "A"
    elif percentage >= 70:
        return "B"
    elif percentage >= 60:
        return "C"
    elif percentage >= 50:
        return "D"
    else:
        return "F"

# STEP 8 - Calculate Student Rank

def calculate_student_totals(ws, subjects):

    student_totals = []

    for row in range(2, ws.max_row + 1):

        student_id = ws.cell(row=row, column=1).value
        student_name = ws.cell(row=row, column=2).value

        total = 0

        for column in subjects.values():

            marks = ws.cell(row=row, column=column).value
            total = total + marks

        student_totals.append({
            "id": student_id,
            "name": student_name,
            "total": total
        })

    return student_totals

def calculate_student_rank(student_totals, student_row, ws):

    target_total = 0

    for student in student_totals:

        if student["id"] == ws.cell(row=student_row, column=1).value:
            target_total = student["total"]
            break

    rank = 1

    for student in student_totals:

        if student["total"] > target_total:
            rank = rank + 1

    return rank

# STEP 9 - Calculate Class Statistics

def calculate_class_statistics(ws, subjects):

    class_statistics = {}

    for subject, column in subjects.items():

        marks_list = []

        for row in range(2, ws.max_row + 1):

            marks = ws.cell(row=row, column=column).value
            marks_list.append(marks)

        average = sum(marks_list) / len(marks_list)
        highest = max(marks_list)
        lowest = min(marks_list)

        class_statistics[subject] = {
            "average": average,
            "highest": highest,
            "lowest": lowest
        }

    return class_statistics

# STEP 10 - Display Class Statistics

def display_class_statistics(ws, subjects):

    class_statistics = calculate_class_statistics(ws, subjects)

    print("\n" + "=" * 60)
    print("                 CLASS STATISTICS")
    print("=" * 60)

    print("\nTotal Students :", ws.max_row - 1)
    print("Total Subjects :", len(subjects))

    print("\n" + "-" * 60)
    print("Subject Statistics")
    print("-" * 60)

    for subject, statistics in class_statistics.items():

        print(f"\n{subject}")
        print(f"Average : {statistics['average']:.2f}")
        print(f"Highest : {statistics['highest']}")
        print(f"Lowest  : {statistics['lowest']}")

    print("\n" + "=" * 60)

# STEP 11 - Display Top Performers

def display_top_performers(ws, subjects):

    student_totals = calculate_student_totals(ws, subjects)

    # Sort students from highest total to lowest
    student_totals.sort(
        key=lambda student: student["total"],
        reverse=True
    )

    print("\n" + "=" * 60)
    print("                    TOP PERFORMERS")
    print("=" * 60)

    print("\n" + "-" * 60)
    print(f"{'Rank':<8}{'Student':<15}{'Total':<12}{'Percentage'}")
    print("-" * 60)

    max_marks = len(subjects) * 100

    for student in student_totals[:5]:

        rank = 1 + sum(
            other_student["total"] > student["total"]
            for other_student in student_totals
        )

        percentage = (student["total"] / max_marks) * 100

        print(
            f"{rank:<8}"
            f"{student['name']:<15}"
            f"{student['total']:<12}"
            f"{percentage:.2f}%"
        )

    print("\n" + "=" * 60)

# STEP 12 - Compare Student With Class Average

def compare_with_class_average(ws, subjects, student_row, student_marks):

    print("\n" + "-" * 60)
    print("SUBJECT-WISE CLASS COMPARISON")
    print("-" * 60)

    print(
        f"{'Subject':<15}"
        f"{'Student':<12}"
        f"{'Class Avg':<12}"
        f"{'Difference'}"
    )

    print("-" * 60)

    for subject, column in subjects.items():

        student_mark = student_marks[subject]

        total_marks = 0

        for row in range(2, ws.max_row + 1):

            marks = ws.cell(
                row=row,
                column=column
            ).value

            total_marks = total_marks + marks

        student_count = ws.max_row - 1

        class_average = total_marks / student_count

        difference = student_mark - class_average

        print(
            f"{subject:<15}"
            f"{student_mark:<12}"
            f"{class_average:<12.2f}"
            f"{difference:+.2f}"
        )

    print("-" * 60)

# STEP 13 - Display Overall Class Performance

def display_class_overview(ws, subjects):

    class_statistics = calculate_class_statistics(ws, subjects)
    student_totals = calculate_student_totals(ws, subjects)

    # Calculate overall class average
    subject_averages = []

    for statistics in class_statistics.values():
        subject_averages.append(statistics["average"])

    overall_average = sum(subject_averages) / len(subject_averages)

    # Find best and lowest performing subjects
    best_subject = max(
        class_statistics,
        key=lambda subject: class_statistics[subject]["average"]
    )

    lowest_subject = min(
        class_statistics,
        key=lambda subject: class_statistics[subject]["average"]
    )

    # Find highest and lowest scoring students
    highest_student = max(
        student_totals,
        key=lambda student: student["total"]
    )

    lowest_student = min(
        student_totals,
        key=lambda student: student["total"]
    )

    print("\n" + "=" * 60)
    print("              CLASS PERFORMANCE OVERVIEW")
    print("=" * 60)

    print("\nTotal Students :", ws.max_row - 1)
    print("Total Subjects :", len(subjects))

    print("\nOverall Class Average :", f"{overall_average:.2f}%")

    print(
        "\nBest Performing Subject :",
        best_subject,
        f"({class_statistics[best_subject]['average']:.2f})"
    )

    print(
        "Lowest Performing Subject:",
        lowest_subject,
        f"({class_statistics[lowest_subject]['average']:.2f})"
    )

    print(
        "\nHighest Scoring Student :",
        highest_student["name"],
        f"({highest_student['total']})"
    )

    print(
        "Lowest Scoring Student  :",
        lowest_student["name"],
        f"({lowest_student['total']})"
    )

    print("\n" + "=" * 60)


# STEP 14 - Analyze Subject-wise Pass/Fail Status

def analyze_subject_results(student_marks, pass_mark=40):

    subject_results = {}

    for subject, marks in student_marks.items():

        if marks >= pass_mark:
            status = "PASS"
        else:
            status = "FAIL"

        subject_results[subject] = {
            "marks": marks,
            "status": status
        }

    return subject_results

# STEP 15 - Generate Complete Student Report

def display_student_report(
        ws,
        student_row,
        student_marks,
        total,
        average,
        percentage,
        rank,
        highest_subject,
        highest_marks,
        lowest_subject,
        lowest_marks,
        performance_level,
        grade,
        subject_results
    ):

    student_id = ws.cell(row=student_row, column=1).value
    student_name = ws.cell(row=student_row, column=2).value

    print("\n" + "=" * 60)
    print("              STUDENT PERFORMANCE REPORT")
    print("=" * 60)

    print("\nStudent ID :", student_id)
    print("Student    :", student_name)

    print("\n" + "-" * 60)
    print("Subject Marks")
    print("-" * 60)

    for subject, marks in student_marks.items():
        status = subject_results[subject]["status"]
        print(f"{subject:<12}: {marks:<8} {status}")
        

     # Calculate overall subject result
    passed_subjects = sum(
        result["status"] == "PASS"
        for result in subject_results.values()
    )

    failed_subjects = sum(
        result["status"] == "FAIL"
        for result in subject_results.values()
    )

    overall_result = "PASS" if failed_subjects == 0 else "FAIL"

    print("\n" + "-" * 60)
    print("Result Summary")
    print("-" * 60)

    print(f"Subjects Passed : {passed_subjects}")
    print(f"Subjects Failed : {failed_subjects}")
    print(f"Result          : {overall_result}")

    print("\n" + "-" * 60)
    print("Performance Summary")
    print("-" * 60)

    max_marks = len(student_marks) * 100

    print(f"Total      : {total} / {max_marks}")
    print(f"Average    : {average:.2f}")
    print(f"Percentage : {percentage:.2f}%")
    print(f"Grade      : {grade}")
    print(f"Rank       : {rank}")

    print("\n" + "-" * 60)
    print("Performance Insights")
    print("-" * 60)

    print(f"Highest Subject : {highest_subject} ({highest_marks})")
    print(f"Lowest Subject  : {lowest_subject} ({lowest_marks})")
    print(f"Performance     : {performance_level}")

    print("\n" + "=" * 60)


# STEP 18 - Export Student Report to Excel

def export_student_report(
        ws,
        student_row,
        student_marks,
        total,
        average,
        percentage,
        rank,
        highest_subject,
        highest_marks,
        lowest_subject,
        lowest_marks,
        performance_level
):

    student_id = ws.cell(row=student_row, column=1).value
    student_name = ws.cell(row=student_row, column=2).value

    # Create a new workbook
    report_wb = Workbook()
    report_ws = report_wb.active

    report_ws.title = "Student Report"

    # Report title
    report_ws["A1"] = "STUDENT PERFORMANCE REPORT"

    # Student details
    report_ws["A3"] = "Student ID"
    report_ws["B3"] = student_id

    report_ws["A4"] = "Student Name"
    report_ws["B4"] = student_name

    # Subject marks
    report_ws["A6"] = "Subject"
    report_ws["B6"] = "Marks"

    row = 7

    for subject, marks in student_marks.items():
        report_ws.cell(row=row, column=1).value = subject
        report_ws.cell(row=row, column=2).value = marks
        row = row + 1

    # Performance summary
    row = row + 1

    report_ws.cell(row=row, column=1).value = "Total"
    report_ws.cell(row=row, column=2).value = total

    row = row + 1

    report_ws.cell(row=row, column=1).value = "Average"
    report_ws.cell(row=row, column=2).value = average

    row = row + 1

    report_ws.cell(row=row, column=1).value = "Percentage"
    report_ws.cell(row=row, column=2).value = percentage

    row = row + 1

    report_ws.cell(row=row, column=1).value = "Rank"
    report_ws.cell(row=row, column=2).value = rank

    # Performance insights
    row = row + 2

    report_ws.cell(row=row, column=1).value = "Highest Subject"
    report_ws.cell(row=row, column=2).value = highest_subject

    row = row + 1

    report_ws.cell(row=row, column=1).value = "Highest Marks"
    report_ws.cell(row=row, column=2).value = highest_marks

    row = row + 1

    report_ws.cell(row=row, column=1).value = "Lowest Subject"
    report_ws.cell(row=row, column=2).value = lowest_subject

    row = row + 1

    report_ws.cell(row=row, column=1).value = "Lowest Marks"
    report_ws.cell(row=row, column=2).value = lowest_marks

    row = row + 1

    report_ws.cell(row=row, column=1).value = "Performance"
    report_ws.cell(row=row, column=2).value = performance_level

    # Adjust column widths
    report_ws.column_dimensions["A"].width = 25
    report_ws.column_dimensions["B"].width = 25

    # Save the report
    filename = f"{student_name}_Report.xlsx"

    report_wb.save(filename)

    print(f"\nReport exported successfully!")
    print(f"File: {filename}")

# STEP 16 - Search Student Function

def search_student(ws, subjects):

    search_value = input("\nEnter Student ID or Name: ").strip()

    student_row = find_student(ws, search_value)

    if student_row is None:
        print("\nStudent not found.")
        return

    print("\nStudent found!")
    print("Excel Row:", student_row)

    # Get student's subject marks
    student_marks = get_student_marks(
        ws,
        student_row,
        subjects
    )

    # Calculate student's performance
    total, average, percentage = calculate_performance(
        student_marks
    )

    # Calculate totals of all students
    student_totals = calculate_student_totals(
        ws,
        subjects
    )

    # Calculate student's rank
    rank = calculate_student_rank(
        student_totals,
        student_row,
        ws
    )

    # Analyze student's performance
    highest_subject, highest_marks, lowest_subject, lowest_marks, performance_level = analyze_student_performance(
        student_marks
    )
    # Calculate student's grade
    grade = calculate_grade(percentage)

    # Analyze subject-wise results
    subject_results = analyze_subject_results(student_marks)


    # Display complete report
    display_student_report(
        ws,
        student_row,
        student_marks,
        total,
        average,
        percentage,
        rank,
        highest_subject,
        highest_marks,
        lowest_subject,
        lowest_marks,
        performance_level,
        grade,
        subject_results
    )
    export_choice = input(
        "\nDo you want to export this report to Excel? (y/n): "
    ).strip().lower()

    if export_choice == "y":

        export_student_report(
            ws,
            student_row,
            student_marks,
            total,
            average,
            percentage,
            rank,
            highest_subject,
            highest_marks,
            lowest_subject,
            lowest_marks,
            performance_level
        )

    compare_with_class_average(
        ws,
        subjects,
        student_row,
        student_marks
    )
# STEP 17 - Main Menu

while True:

    print("\n" + "=" * 60)
    print("        STUDENT PERFORMANCE ANALYZER")
    print("=" * 60)

    print("1. Search Student")
    print("2. View Class Statistics")
    print("3. View Top Performers")
    print("4. Exit")

    choice = input("\nEnter your choice: ").strip()

    if choice == "1":
        search_student(ws, subjects)

    elif choice == "2":
        display_class_statistics(ws, subjects)
        display_class_overview(ws, subjects)

    elif choice == "3":
        display_top_performers(ws, subjects)

    elif choice == "4":
        print("\nThank you for using Student Performance Analyzer!")
        break

    else:
        print("\nInvalid choice. Please enter 1, 2, 3 or 4.")

        # OUTPUT:
        
        # Do you want to export this report to Excel? (y/n): y

        # Report exported successfully!
        # File: Anjali_Report.xlsx





