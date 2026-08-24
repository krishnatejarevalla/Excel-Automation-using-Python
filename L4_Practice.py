from openpyxl import load_workbook
from openpyxl.utils import range_boundaries, column_index_from_string
wb = load_workbook("MarkList.xlsx")
ws = wb.active

#Printing the boundaries of a given range
cell_range = input("Enter the range of cells: ")
min_col, min_row, max_col, max_row = range_boundaries(cell_range)
print(f"Starting row : {min_row}",
       f"Starting col : {min_col}",
       f"ending row :{max_row}",
       f"ending col : {max_col}"
       , end =",")
        #Output:
        #Enter the range of cells: A1:F11
        #Starting row : 1 Starting col : 1 ending row :11 ending col : 6,

#printing cell values from given range
for row in range(min_row, max_row+1):
     for col in range(min_col, max_col+1):
        print(ws.cell(row = row, column= col).value, end = " ")
    print()
                # Output:
                # Enter the range of cells: A1:F11
                # Starting row : 1 Starting col : 1 ending row :11 ending col : 6,Student_ID Name Python SQL Excel PowerBI 
                # 101 Ravi 78 85 90 82 
                # 102 Priya 92 88 84 91 
                # 103 Arjun 65 72 70 68 
                # 104 Sneha 88 94 91 89 
                # 105 Kiran 74 69 80 76 
                # 106 Anjali 95 91 96 94 
                # 107 Rahul 81 77 85 79 
                # 108 Divya 89 93 88 90 
                # 109 Vijay 70 68 75 72 
                # 110 Neha 86 90 92 87 


#Printing values from specific column
Column_letter = input("Enter column letter: ")
col_num = column_index_from_string(Column_letter)
for row in range(min_row, max_row+1):
    print(ws.cell(row=row, column= col_num).value)

#Searching for a value and it's cell coordinates
Keyword = input("Enter the value to search : ")
found = False
for row in range(min_row, max_row+1):
    for col in range(min_col, max_col+1):
        if ws.cell(row=row, column = col).value == Keyword:
            print(f"{Keyword} is found at" ,ws.cell(row, col).coordinate)
            found =True
if found is not True:
    print("Not Found")
            #output:
            # Enter the range of cells: A1:F11
            # Starting row : 1 Starting col : 1 ending row :11 ending col : 6,Enter the value to search : Priya
            # Priya is found at B3

