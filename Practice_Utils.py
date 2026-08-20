from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
wb = load_workbook("MarkList.xlsx")
ws = wb.active

#printing column names using get_column_letter
# for col in range(1,7):
#     print("Column "+ str(col) , "=" ,get_column_letter(col))
#             #Output:  Column 1 = A
#                     # Column 2 = B
#                     # Column 3 = C
#                     # Column 4 = D
#                     # Column 5 = E
#                     # Column 6 = F

#Using ws.max_column and ws.cell().value functions
for col in range(1, ws.max_column+1):
    print(get_column_letter(col)+"=",ws.cell(row=1,column = col).value)
            # OUTPUT:
            # A= Student_ID
            # B= Name
            # C= Python
            # D= SQL
            # E= Excel
            # F= PowerBI

#Using column_index_from_string
from openpyxl.utils import column_index_from_string
for letter in range(65, 65 + ws.max_column):
    letter = chr(letter)
    print(letter, "=", column_index_from_string(letter))
                # OUTPUT
                # A = 1
                # B = 2
                # C = 3
                # D = 4
                # E = 5
                # F = 6 