#Automating Excel using openpyxl module

#Accessing Worksheets from workbooks and cellvalues present in it
from openpyxl import Workbook, load_workbook
wb = load_workbook("Marks.xlsx")
ws = wb.active
print(ws['A2'].value) #Arjun
print(ws['A3'].value) #Priya
print(ws['B6'].value) #69
print(ws.max_row) #9 rows
print(ws.max_column) #5 cols

#changing the cellvalues
ws['A2'] = "Teja"
ws['A3'] = "Srinu"
ws['B6'] = 79

#saving the changes into Excel file
wb.save('Marks.xlsx')


#Accessing the sheet names
print(wb.sheetnames) #Grades

#Creating new sheets in the wb
wb.create_sheet('Averages')
wb.create_sheet('Remarks')
wb.create_sheet('Student Info')
print(wb.sheetnames)
