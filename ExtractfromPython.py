#Extracting the data from the Python file into an Excel Worksheet
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

data = {
	"Joe": {
		"math": 65,
		"science": 78,
		"english": 98,
		"gym": 89
	},
	"Bill": {
		"math": 55,
		"science": 72,
		"english": 87,
		"gym": 95
	},
	"Tim": {
		"math": 100,
		"science": 45,
		"english": 75,
		"gym": 92
	},
	"Sally": {
		"math": 30,
		"science": 25,
		"english": 45,
		"gym": 100
	},
	"Jane": {
		"math": 100,
		"science": 100,
		"english": 100,
		"gym": 60
	}
}

#Placing the above data into an Excel Sheet
wb = Workbook()
ws = wb.active
ws.title = "Scores"

#To Name the column headings, I'm using list[] and ws.append()
headings = ['Name'] + list(data["Joe"].keys())
ws.append(headings) #This Gives column headers |Name|math|science|english|gym| in excel


#Now, we've to get Student names and their scores, so I'm using for loop
for person in data:
    grades = list(data[person].values()) #make sure to convert dict to list values
    ws.append([person]+grades)
    
wb.save("MarkSheet.xlsx")