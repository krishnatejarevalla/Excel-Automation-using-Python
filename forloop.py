#using for loop for accessing cellvalues
from openpyxl import workbook, load_workbook
from openpyxl.utils import get_column_letter
wb = load_workbook('Marks.xlsx')
ws = wb.active
for row in range(1,7):
    for col in range(1,6):
        # char = chr(65 + col)
        # print(ws[char+str(row)])
        char = get_column_letter(col)
        print(ws[char+str(row)].value)