# =========================================
#      Student Performance Analyser
# =========================================


# STEP 1 — Loading the WorkBook
from openpyxl import load_workbook
wb = load_workbook("MarkList.xlsx")
ws = wb.active
print("===========================================================")
print("===================Excel Mini Project======================")
print("===========================================================")
print("Rows:",ws.max_row)
print("Columns:",ws.max_column)

# STEP 2 — Understand which columns are subjects
print("\nColumn Headers:")
for col in range(1, ws.max_column+1):
    print(col, ws.cell(row=1, column =col).value)
        #Output:
        # Column Headers:
        # 1 Student_ID
        # 2 Name
        # 3 Python
        # 4 SQL
        # 5 Excel
        # 6 PowerBI

# STEP 3 — Identify the subject columns
subject_start = 3
subject_end = ws.max_column
print("\nSubjects:")
for col in range(subject_start, subject_end+1):
    print(ws.cell(row=1,column=col).value)


# STEP 4 — Calculate TOTAL for each student and Percentage
print("\nStudent Total Marks")
for row in range(2, ws.max_row+1):
    total = 0
    for col in range(subject_start, subject_end+1):
        marks = ws.cell(row = row, column = col).value
        total = total + marks
    max_marks = (subject_end - subject_start +1)*100
    percentage = (total/max_marks)*100
    print(ws.cell(row=row,column=2).value,
          "Total:",total,
          "Percentage:",percentage)

        #Output:
        # Student Total Marks
        # Ravi Total: 335 Percentage: 83.75
        # Priya Total: 355 Percentage: 88.75
        # Arjun Total: 275 Percentage: 68.75
        # Sneha Total: 362 Percentage: 90.5
        # Kiran Total: 299 Percentage: 74.75
        # Anjali Total: 376 Percentage: 94.0
        # Rahul Total: 322 Percentage: 80.5
        # Divya Total: 360 Percentage: 90.0
        # Vijay Total: 285 Percentage: 71.25
        # Neha Total: 355 Percentage: 88.75

# STEP 5 - Calculate AVERAGE for each subject
for col in range(subject_start, subject_end+1):
    total_marks = 0
    for row in range(2, ws.max_row+1):
        marks = ws.cell(row=row,column=col).value
        total_marks = marks+total_marks
    student_count = ws.max_row-1
    average = total_marks/student_count
    print(ws.cell(row=1,column=col).value,"Average:",average)
        #Output:
        # Python Average: 81.8
        # SQL Average: 82.7
        # Excel Average: 85.1
        # PowerBI Average: 82.8